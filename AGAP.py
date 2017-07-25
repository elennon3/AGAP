from __future__ import print_function

# Import Python wrapper for or-tools constraint solver.
from ortools.constraint_solver import pywrapcp

import random
import openpyxl
import csv

def main():
  # Create the solver.
  solver = pywrapcp.Solver('jobshop')

  wb = openpyxl.load_workbook('aircraft_specs.xlsx')
  sheet1 = wb.get_sheet_by_name('Sheet1')

  wb = openpyxl.load_workbook('schedule.xlsm')
  sheet3 = wb.get_sheet_by_name('Sheet1')

  # Number of determined flight pairings with time difference less than 3 hours
  NumPairs = 104

  # Number of gates at airport (Deduct one as iterations begin at zero)
  NumGates = 36

  # Declare lists to act as databases for import from 'aircraft_specs'
  aircraft_db = []
  wingspan_db = []
  length_db = []
  passengers_db = []
  airline_db = []
  lf_db = []

  # Declare international, domestic, Ireland, Channel Islands for random assignment
  nature_db = ['I', 'D', 'IRL', 'CH']

  # Declare lists for aircraft based on range specifications
  aircraft_LH = []
  aircraft_MH = []
  aircraft_SH = []
  aircraft = []

  # Data to generate for each flight to reduce available gates
  wingspan = []
  length = []
  nature = []
  passengers = []
  airline = []
  load_factor = []

  # Import aircraft specification data from spreadsheet
  for i in range(2, 22, 1):
    ac = sheet1.cell(row=i, column=1).value
    ws = sheet1.cell(row=i, column=3).value
    lgth = sheet1.cell(row=i, column=4).value
    pax = sheet1.cell(row=i, column=2).value
    arl = sheet3.cell(row=i, column=1).value
    lf = sheet3.cell(row=i, column=2).value

    # Fill databases
    aircraft_db.extend([[ac]])
    wingspan_db.extend([[ws]])
    length_db.extend([[lgth]])
    passengers_db.extend([[pax]])
    airline_db.extend([[arl]])
    lf_db.extend([[lf]])

    # Divide aircraft into short, medium and long haul
    if sheet1.cell(row=i, column=8).value == 'LH':
      ac_LH = sheet1.cell(row=i, column=1).value
      aircraft_LH.extend([[ac_LH]])

    elif sheet1.cell(row=i, column=8).value == 'MH':
      ac_MH = sheet1.cell(row=i, column=1).value
      aircraft_MH.extend([[ac_MH]])

    elif sheet1.cell(row=i, column=8).value == 'SH':
      ac_SH = sheet1.cell(row=i, column=1).value
      aircraft_SH.extend([[ac_SH]])

  # Print to test if data was gathered correctly
  print("aircraft:\t\t\t", aircraft_db)
  print("Short Haul: \t\t", aircraft_SH)
  print("Medium Haul: \t\t", aircraft_MH)
  print("Long Haul: \t\t\t", aircraft_LH)
  print("Wingspan (m):\t\t", wingspan_db)
  print("Length (m): \t\t", length_db)
  print("Passengers: \t\t", passengers_db)
  print("Airline: \t\t\t", airline_db)
  print("Load Factor: \t\t", lf_db)

  # Assign aircraft to a flight based on its destination
  for i in range(0, NumPairs, 1):
    nat = random.choice(nature_db)
    nature.extend([[nat]])

    if nat == 'I':
      ac = random.choice(aircraft_LH)
      aircraft.extend([ac])

    elif nat == 'D' or nat == 'CH':
      ac = random.choice(aircraft_SH)
      aircraft.extend([ac])

    elif nat == 'IRL':
      ac = random.choice(aircraft_MH)
      aircraft.extend([ac])

    x = aircraft_db.index(ac)
    ws = wingspan_db[x][0]
    lgth = length_db[x][0]

    #assign an airline to the flight to get load factor
    arl = random.choice(airline_db)
    y = airline_db.index(arl)
    lf = lf_db[y][0]

    #calculate passengers on board from aircraft type and load factor
    pax = lf*passengers_db[x][0]


    #add all information synchronously for each flight
    wingspan.extend([[ws]])
    length.extend([[lgth]])
    passengers.extend([[pax]])
    load_factor.extend([[lf]])
    airline.extend([arl])

  #open spreadsheet to read gate specific information
  wb = openpyxl.load_workbook('gate_info.xlsx')
  sheet2 = wb.get_sheet_by_name('gates')

  wb = openpyxl.load_workbook('walking_distances.xlsx')
  sheet4 = wb.get_sheet_by_name('Security')

  # Define gates list and data
  gate_number = []
  gate_length_db = []
  gate_nature = []
  gate_clearance_db = []
  walking_distance_db = []

  for i in range(2, 38, 1):
    ga = sheet2.cell(row=i, column=1).value
    lgth = sheet2.cell(row=i, column=5).value
    nat = sheet2.cell(row=i, column=4).value
    gate_clrnc = sheet2.cell(row=i, column=6).value
    walk_dist = sheet4.cell(row=i, column=3).value

    gate_number.extend([[ga]])
    gate_length_db.extend([[lgth]])
    gate_nature.extend([[nat]])
    gate_clearance_db.extend([[gate_clrnc]])
    walking_distance_db.extend([[walk_dist]])


  ###########################################################################################
  #               END OF LIST CREATION
  ###########################################################################################
  gates = []
  park_time = []
  arrival_time = []
  gate_length_test = []
  gate_clearance = []
  walking_distance = []
  walking_distance1 = []
  gates1 = []

  # Generate gate sets which are suitable for flight type at Glasgow airport
  for i in range(0, NumPairs, 1):

    nat=nature[i][0]
    gate_set=[]
    gates1=[]
    count=0

    if nat == 'I':
      for k in range(27,37,1):
        gate_set.extend([[k]])
        count=count+1

    elif nat == 'D':
      for k in range(14,27,1):
        gate_set.extend([[k]])
        count = count + 1

    elif nat == 'IRL' or nat == 'CH':
      for k in range(0,13,1):
        gate_set.extend([[k]])
        count = count + 1

    for x in range(0,count,1):

      gate = gate_set[x][0]

      j = gate_number.index([gate])
      walk_dist = walking_distance_db[j][0]
      gate_lgth = gate_length_db[j][0]
      gate_clrnc = gate_clearance_db[j][0]
      ac_lgth = length[i][0]
      wgspan = wingspan[i][0]

      #Assign gate based on aircraft length
      if ac_lgth < gate_lgth and wgspan < gate_clrnc:

        gates1.extend([gate])
        walking_distance1.extend([walk_dist])

    gate = random.choice(gates1)
    j = gate_number.index([gate])
    walk_dist = walking_distance_db[j][0]
    gate_lgth = gate_length_db[j][0]
    gate_clrnc = gate_clearance_db[j][0]
    ac_lgth = length[i][0]
    wgspan = wingspan[i][0]

    print("gates1: \t\t\t", gates1)
    print("ac length: \t\t", ac_lgth)
    print("gate length: \t", gate_lgth)
    print("wingspan: \t\t", wgspan)
    print("gate clrnc: \t", gate_clrnc)
    print("walk dist: \t\t", walk_dist)
    print("iter gate: \t\t", gate)

    gates.extend([[gate]])
    gate_length_test.extend([gate_lgth])
    gate_clearance.extend([gate_clrnc])
    print("SELECT GATE: \t\t", gate)

    #Random gate-in time for each flight (rounded to five mins)
    park_time.extend([[random.randint(4, 20) * 5]])

    #Random arrival time for flight pair (mins from zero to 24 hours)
    arrival_time.extend([[random.randint(0, 1440)]])

    # convert arrival time to decimal hours for printing
    arrival_time_hours = [[x[0] / float(60)] for x in arrival_time]

  print("\n\nGate:\t\t\t\t\t", gates)
  print("Gate-in times (mins):\t", park_time)
  print("Arrival times (mins):\t", arrival_time)
  print("Aircraft:\t\t\t\t", aircraft)
  print("Wingspan:\t\t\t\t", wingspan)
  print("Gate clearance:\t\t\t", gate_clearance)
  print("Aircraft length:\t\t", length)
  print("Gate length:\t\t\t", gate_length_test)
  print("Passengers:\t\t\t\t", passengers)
  print("Nature:\t\t\t\t\t", nature)
  print("Airline:\t\t\t\t", airline)
  print("Load factor:\t\t\t", load_factor)

  gates_count = NumGates
  pair_count = NumPairs
  all_gates = range(0, gates_count)
  all_pairs = range(0, pair_count)

  # Computes horizon.
  horizon = 0
  for i in all_gates:
    horizon += sum(park_time[i])
  # Creates jobs.
  all_tasks = {}
  for i in all_pairs:
    for j in range(0, len(gates[i])):
      all_tasks[(i, j)] = solver.FixedDurationIntervalVar(0,
                                                          horizon,
                                                          park_time[i][j],
                                                          False,
                                                          '\tPair_%i_%i \t' % (i, j))

  # Creates sequence variables and add disjunctive constraints.
  all_sequences = []
  all_machines_jobs = []
  for i in all_gates:

    gates_jobs = []
    for j in all_pairs:
      for k in range(0, len(gates[j])):
        if gates[j][k] == i:
          gates_jobs.append(all_tasks[(j, k)])
    disj = solver.DisjunctiveConstraint(gates_jobs, 'gate %i' % i)
    all_sequences.append(disj.SequenceVar())
    solver.Add(disj)

  # Add conjunctive contraints.
  for i in all_pairs:
    for j in range(0, len(gates[i]) - 1):
      solver.Add(all_tasks[(i, j + 1)].StartsAfterEnd(all_tasks[(i, j)]))

  # Set the objective.
  obj_var = solver.Max([all_tasks[(i, len(gates[i]) - 1)].EndExpr()
                        for i in all_pairs])
  objective_monitor = solver.Minimize(obj_var, 1)
  # Create search phases.
  sequence_phase = solver.Phase([all_sequences[i] for i in all_gates],
                                solver.SEQUENCE_DEFAULT)
  vars_phase = solver.Phase([obj_var],
                            solver.CHOOSE_FIRST_UNBOUND,
                            solver.ASSIGN_MIN_VALUE)
  main_phase = solver.Compose([sequence_phase, vars_phase])
  # Create the solution collector.
  collector = solver.LastSolutionCollector()

  # Add the interesting variables to the SolutionCollector.
  collector.Add(all_sequences)
  collector.AddObjective(obj_var)

  for i in all_gates:
    sequence = all_sequences[i];
    sequence_count = sequence.Size();
    for j in range(0, sequence_count):
      t = sequence.Interval(j)
      collector.Add(t.StartExpr().Var())
      collector.Add(t.EndExpr().Var())
  # Solve the problem.
  disp_col_width = 10

  if solver.Solve(main_phase, [objective_monitor, collector]):
    print("\nOptimal Schedule Length:", collector.ObjectiveValue(0), "\n")
    sol_line = ""
    sol_line_tasks = ""
    print("Optimal Schedule", "\n")

    for i in all_gates:
      seq = all_sequences[i]
      sol_line += "Gate " + str(i) + ": "
      sol_line_tasks += "Gate " + str(i) + ": "
      sequence = collector.ForwardSequence(0, seq)
      seq_size = len(sequence)

      for j in range(0, seq_size):
        t = seq.Interval(sequence[j]);
        # Add spaces to output to align columns.
        sol_line_tasks += t.Name() + " " * (disp_col_width - len(t.Name()))

      for j in range(0, seq_size):
        t = seq.Interval(sequence[j]);
        sol_tmp = "[" + str(collector.Value(0, t.StartExpr().Var())) + ","
        sol_tmp += str(collector.Value(0, t.EndExpr().Var())) + "] "
        # Add spaces to output to align columns.
        sol_line += sol_tmp + " " * (disp_col_width - len(sol_tmp))

      sol_line += "\n"
      sol_line_tasks += "\n"

    print(sol_line_tasks)
    print("Time Intervals for Tasks\n")
    print(sol_line)


if __name__ == '__main__':
  main()
