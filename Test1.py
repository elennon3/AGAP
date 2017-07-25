from __future__ import print_function
import sys
from ortools.constraint_solver import pywrapcp
import numpy as np
import csv
import openpyxl
import Tkinter as tk
import tkMessageBox
import matplotlib.pyplot as plt
from numpy import array

###########################################################################################
#                       DEFINE FUNCTIONS
##########################################################################################

#Creates csv file of flight pairs by reading data from Excel file, 'Schedule'
def flight_pairs():
    #Open flight schedule workbook with arrivals and departures sheets
    wb = openpyxl.load_workbook('schedule.xlsm')
    sheet1 = wb.get_sheet_by_name('departures')
    sheet2 = wb.get_sheet_by_name('arrivals')
    sheet3 = wb.get_sheet_by_name('Sheet1')

    #initialise pairs of flights
    pairs = 0

    #write flight pair output to csv file
    with open('flight_pairs.csv', 'w') as f1:
        writer = csv.writer(f1, delimiter='\t', lineterminator='\n', )
        writer.writerow(['pair','tail_number', 'arrival_flight_number', 'airline', 'load_factor', 'aircraft', 'passengers', 'wingspan', 'length', 'arrival_time', 'gate_in_time_mins', 'departure_flight_number', 'departure_time', 'dest'])

        #departures iteration
        for i in range(2, 152, 1):
            tail_number_D = sheet1.cell(row=i, column=27).value
            flight_number_D = sheet1.cell(row=i, column=2).value
            destination_D = sheet1.cell(row=i, column=12).value
            aircraft_D = sheet1.cell(row=i, column=3).value
            airline_D = sheet1.cell(row=i, column=13).value
            load_factor = sheet1.cell(row=i, column=30).value
            wingspan = sheet1.cell(row=i, column=22).value
            length = sheet1.cell(row=i, column=23).value
            passengers = sheet1.cell(row=i, column=21).value
            dest = sheet1.cell(row=i, column=28).value


            gate_scheduled_D = int(sheet1.cell(row=i, column=29).value*24*60)

        #arrivals iteration
            for j in range(2, 148, 1):

                flight_number_A = sheet2.cell(row=j, column=2).value
                tail_number_A = sheet2.cell(row=j, column=27).value
                gate_scheduled_A = int(sheet2.cell(row=j, column=29).value*24*60)
                airline_A = sheet2.cell(row=j, column=13).value

                # Create flight pairs based on tail number.
                # Condition 2 ensures aircraft arrived before subsequent departure
                if tail_number_D == tail_number_A and gate_scheduled_A < gate_scheduled_D:

                    #Scheduled time at gate between arrival and departure (minutes)
                    duration = (gate_scheduled_D - gate_scheduled_A)

                    #Limit to gate-in times less than three hours
                    if duration < 180:
                        pairs += 1

                        #write pairing data to csv file declared above
                        row = [pairs, tail_number_A, flight_number_A, airline_A, load_factor, aircraft_D, passengers, wingspan, length, gate_scheduled_A, duration, flight_number_D, gate_scheduled_D, dest]
                        writer.writerow(row)

# Checks which gates have been used in the allocation (PROBABLY WON'T GET USED)
def check_gates(gt):
    gates = []

    for i in range(0, len(gt), 1):
        if gt[i][1] in gates:

            continue

        else:
            gates.append(gt[i][1])

    print('Number of gates used: \t', len(gates))
    print('Gates used: \t\t\t', gates)

#Generates gate destinations for which each gate can be used (CAN BE IMPROVED)
def gate_destination():

    gate_dest = []

    for x in csv.DictReader(open('gate_info.csv'), delimiter='\t'):
        y = []
        z = [0,0,0,0]

        #Reads if gate is suitable for all four destination types
        y.append(int(x['I']))
        y.append(int(x['D']))
        y.append(int(x['IRL']))
        y.append(int(x['CH']))

        #Reorganises gate destinations for each gate k
        i=0
        if y[0]==1:
            z[i]='I'
            i=i+1

        if y[1]==1:
            z[i]='D'
            i=i+1

        if y[2]==1:
            z[i]='IRL'
            i=i+1

        if y[3]==1:
            z[i]='CH'

        gate_dest.append(z)

    #Return list of gate destinations to main()
    return gate_dest

# Reads and returns lists of data corresponding to flight schedule and airport data
def schedule_data():

    #Can be used to reduce number of flights for test purposes
    limit_flights = 5

    #Create warning message that not all flihgt data is included in the simulation
    if 'limit_flights' in locals():
        print("\n\nWARNING: Not all flight data included in simulation\n\n")
        root = tk.Tk()
        root.withdraw()
        tkMessageBox.showwarning('Warning', 'Not all flight data included in simulation. Remove "limit_flights" from schedule_data() to retrieve all flights')

    #Define all lists for data reading

    a = []                  # Arrival time of each flight pair i
    d = []                  # Departure time of each flight pair i
    k = []                  # Gate k
    w = []                  # Walking distance (metres) from security/baggage to gate k
    f = []                  # Estimated arriving and departing passengers on flight pair i
    lf = []                 # Airline passenger load factor (percentage)
    pax = []                # Max number of passengers for aircraft type
    gate_clearance = []     # Maximum wingspan accommodated by gate
    gate_length = []        # Maximum aircraft length accommodated by gate
    airline = []            # Airline operating flight pair i
    aircraft = []           # Aircraft type operating flight pair i
    wingspan = []           # Aircraft wingspan (metres)
    length = []             # Aircraft length (metres)
    dest = []               # Aircraft destination (International, Domestic, Ireland, Channel Islands)
    gate_dest = []          # Gate allowable destinations (International, Domestic, Ireland, Channel Islands)

    #Reads schedule data
    count = 0
    for x in csv.DictReader(open('flight_pairs.csv'), delimiter='\t'):
        a.append(int(x['arrival_time']))
        d.append(int(x['departure_time']))
        lf.append(float(x['load_factor']))
        pax.append(int(x['passengers']))
        wingspan.append(float(x['wingspan']))
        length.append(float(x['length']))
        airline.append(x['airline'])
        aircraft.append(x['aircraft'])
        dest.append(x['dest'])

        #Reduces number of flights for testing purposes
        if 'limit_flights' in locals():
            if count == limit_flights:
                break

            else:
                count = count + 1

    #Reads airline load factor and aircraft passenger numbers
    count=0
    for x in range(0, len(a), 1):
        load_factor = lf[x]
        passengers = pax[x]
        num_pax = load_factor*passengers
        f.append(int(num_pax))

        if 'limit_flights' in locals():
            if count == limit_flights:
                break

            else:
                count = count + 1

    #Read gate information
    for x in csv.DictReader(open('gate_info.csv'), delimiter='\t'):
        k.append(int(x['gate_number']))
        gate_clearance.append(float(x['wingspan']))
        gate_length.append(float(x['max_aircraft_length_m']))


    gate_dest = gate_destination()

    #Reads walking distances to each gate
    for x in csv.DictReader(open('Security_Distances.csv'), delimiter='\t'):
        w.append(int(x['Distance']))

    #Print each dataset
    print('Arrivals = \t\t\t', a)
    print('Departures = \t\t', d)
    print('Airline = \t\t\t', airline)
    print('Aircraft = \t\t\t', aircraft)
    print('Wingspan = \t\t\t', wingspan)
    print('Length = \t\t\t', length)
    print('Gate Numbers = \t\t', k)
    print('Gate Dest. = \t\t', gate_dest)
    print('Walking dist = \t\t', w)
    print('Load factor = \t\t', lf)
    print('Passengers = \t\t', pax)
    print('Est. passengers = \t', f)
    print('Destination = \t\t', dest)
    print('Gate Clearance = \t', gate_clearance)
    print('Gate Length = \t\t', gate_length)

    #Converts separate lists of data into grouped data per flight pair i
    flights = np.column_stack((a,d,f,airline,dest,aircraft,wingspan,length))

    #Outputs grouped data per flight pair i
    print('\n\n Grouped Flight Data:\n')
    print('  Arr.  Dep.  Pax\tAirline\t\tDest.\tAircraft  Wingspan\tLength')
    print(flights)
    print('\n\n')

    #Converts separate lists of data into grouped data per gate k
    gates = np.column_stack((k,w,gate_dest,gate_clearance,gate_length))

    #Outpute grouped data per gate k
    print('\n\n Grouped Gate Data:\n')
    print('  Gate  Walk Dist.  Dest  Clearance Length')
    print(gates)
    print('\n\n')


    #Set of flight pairs arriving at or departing from airport
    N = len(a)
    n = len(a)

    #Set of gates available at airport
    M = len(k)
    m = len(k)

    #Initialise available gate times
    g = [-1]*m

    #Return gathered data to main()
    return a,d,k,w,gate_clearance,gate_length,airline,aircraft,wingspan,length,lf,pax,f,dest,g,N,n,M,m,gates,flights


def assign():
    #Set limit on number of output solutions for testing purposes
    solution_limit = 1

    #Create warning message that not all solutions will be displayed
    if 'solution_limit' in locals():
        root = tk.Tk()
        root.withdraw()
        tkMessageBox.showwarning('Warning','Not all solutions will be displayed. Remove "solution_limit" from main() to find all solutions')

    # Creates the AGAP solver
    solver = pywrapcp.Solver("AGAP")


    #Generates flight and gate info from schedule_data()
    a, d, k, w, gate_clearance, gate_length, airline, aircraft, wingspan, length, lf, pax, f, dest, g, N, n, M, m, gates, flights = schedule_data()

    # Creates the variables.
    # The array index is the column, and the value is the row.
    # Rows represent individual gates, and columns represent flight pairs.
    # bin_assign is the binary assignment matrix for flights to gates.
    y = [solver.IntVar(0, N-1, "x%i" % i) for i in range(N)]
    print(y, "\n\n")
    print(y[4])

    wingspan = [solver.IntVar(0, 0, "x%i" %i) for i in range(N)]
    gate_clearance = [solver.IntVar(0, 0, "x%i" % i) for i in range(N)]
    print(wingspan, "\n\n")
    print(gate_clearance, "\n\n")

    #ADD CONSTRAINTS (Blocked out as need to be implemented separately)
    solver.Add(solver.AllDifferent(y))
    #for k in range(N):
        #for i in range(N):
            #for j in range(i):
                #solver.Add(y[i][k] * y[j][k] * (d[j] - a[i]) * (d[i] - a[i]) <= 0)
                #print(i,j,k)

    # for i in range(n):
    #     for j in range(i):
    #         solver.Add(abs(q[i]-q[j]) != abs(i-j))

    # symmetry breaking
    # solver.Add(q[0] == 0)

    #
    # solution and search
    #
    solution = solver.Assignment()
    solution.Add([y[i] for i in range(N)])

    collector = solver.AllSolutionCollector(solution)
    # collector = solver.FirstSolutionCollector(solution)
    # search_log = solver.SearchLog(100, x[0])
    solver.Solve(solver.Phase([y[i] for i in range(N)],
                              solver.INT_VAR_SIMPLE,
                              solver.ASSIGN_MIN_VALUE),
                 [collector])

    num_solutions = collector.SolutionCount()
    print("num_solutions: ", num_solutions)
    if num_solutions > 0:
        for s in range(num_solutions):
            gate_assign = [collector.Value(s, y[i]) for i in range(N)]
            print(gate_assign)
            print("Gate Assign:", gate_assign)
            for i in range(N):
                for j in range(N):
                    if gate_assign[i] == j:
                        print("1", end=' ')
                    else:
                        print("0", end=' ')
                print()
            print()

            if 'solution_limit' in locals():
                break

        print()
        print("num_solutions:", num_solutions)
        print("failures:", solver.Failures())
        print("branches:", solver.Branches())
        print("WallTime:", solver.WallTime())

    else:
        print("No solutions found")

#Main script
flight_pairs()
assign()
