import openpyxl
import csv

#Open flight schedule workbook with arrivals and departures sheets
wb = openpyxl.load_workbook('schedule.xlsm')
sheet1 = wb.get_sheet_by_name('departures')
sheet2 = wb.get_sheet_by_name('arrivals')
sheet3 = wb.get_sheet_by_name('Sheet1')

#initialise pairs of flights
pairs = 0

#write flight pair output to csv file
with open('flight_pairs.csv', 'w') as f1:
    writer = csv.writer(f1, delimiter='\t', lineterminator='\n', )
    writer.writerow(['pair','tail_number', 'arrival_flight_number', 'airline', 'load_factor', 'aircraft', 'passengers', 'wingspan', 'length', 'arrival_time', 'gate_in_time_mins', 'departure_flight_number', 'departure_time', 'dest'])

    #departures iteration
    for i in range(2, 152, 1):
        tail_number_D = sheet1.cell(row=i, column=27).value
        flight_number_D = sheet1.cell(row=i, column=2).value
        destination_D = sheet1.cell(row=i, column=12).value
        aircraft_D = sheet1.cell(row=i, column=3).value
        airline_D = sheet1.cell(row=i, column=13).value
        load_factor = sheet1.cell(row=i, column=30).value
        wingspan = sheet1.cell(row=i, column=22).value
        length = sheet1.cell(row=i, column=23).value
        passengers = sheet1.cell(row=i, column=21).value
        dest = sheet1.cell(row=i, column=28).value


        gate_scheduled_D = int(sheet1.cell(row=i, column=29).value*24*60)

    #arrivals iteration
        for j in range(2, 148, 1):

            flight_number_A = sheet2.cell(row=j, column=2).value
            tail_number_A = sheet2.cell(row=j, column=27).value
            gate_scheduled_A = int(sheet2.cell(row=j, column=29).value*24*60)
            airline_A = sheet2.cell(row=j, column=13).value

            # Create flight pairs based on tail number.
            # Condition 2 ensures aircraft arrived before subsequent departure
            if tail_number_D == tail_number_A and gate_scheduled_A < gate_scheduled_D:

                #Scheduled time at gate between arrival and departure (minutes)
                duration = (gate_scheduled_D - gate_scheduled_A)

                #Limit to gate-in times less than three hours
                if duration < 180:
                    print(flight_number_D + ' ' + flight_number_A)
                    print(duration)
                    pairs += 1

                    #write pairing data to csv file declared above
                    row = [pairs, tail_number_A, flight_number_A, airline_A, load_factor, aircraft_D, passengers, wingspan, length, gate_scheduled_A, duration, flight_number_D, gate_scheduled_D, dest]
                    writer.writerow(row)

print(pairs)